# 起别名
import openpyxl as x
# 这个是图标包
from openpyxl.chart import BarChart,Reference 
# 这里获取文件是从Py根目录搜索
wb= x.load_workbook("py_test.xlsx") # 表格文件名
Sheet=wb['Sheet1'] #Xl文件中的表单名
cell=Sheet.cell(1,1) # 等价于sheet['a1']

print("\n")
print(cell.value) 
print(cell) 
print(Sheet.max_column)
print(Sheet.max_row)
print("\n")

cell=Sheet.cell(1,4)
cell.value="乘0.9的年龄"
for row in range(2,Sheet.max_row+1):
    #这里的3是我们习惯是用的3，不是从0开始
    cell=Sheet.cell(row,3)
    #计算新的年龄
    corrected_price=cell.value*0.9
    #创建新的cell在第四列
    corrected_price_cell=Sheet.cell(row,4)
    #真正赋值
    corrected_price_cell.value=corrected_price
#创建Reference，保存你需要的值
values= Reference(Sheet
          ,min_row=2
          ,max_row=Sheet.max_row
          ,min_col=4
          ,max_col=4)
chart=BarChart()
chart.add_data(values)
Sheet.add_chart(chart,'a5')





#这里创建新的xlxs，而不是覆盖原来的
wb.save("py_new.xlsx")